"""Stand-alone export builders backed by the deterministic domain fixtures."""

from __future__ import annotations

from copy import copy
from html import escape
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..domain.services import (
    class_analytics,
    fixture_exam,
    knowledge_nodes,
    prerequisite_path,
    seeded_attempts,
    seeded_students,
    student_report,
)

SCHOOL_NAME = "โรงเรียนสาธิต TeacherOS"
PROTOTYPE_NOTE = "เอกสารต้นแบบเพื่อการสาธิตระบบ TeacherOS — ไม่ใช่เอกสารราชการ"
BLUE = "0757D9"
TEAL = "0E9F9A"
PALE_BLUE = "EAF2FF"
PALE_TEAL = "E8F7F5"
THIN_GREY = Side(style="thin", color="CBD5E1")


def _e(value: object) -> str:
    """Escape a dynamic value before inserting it into trusted document markup."""
    return escape(str(value), quote=True)


def _one_decimal(value: object) -> str:
    return f"{float(value):.1f}"


def _grade(percentage: float) -> str:
    for threshold, grade in ((80, "4"), (75, "3.5"), (70, "3"), (65, "2.5"), (60, "2"), (55, "1.5"), (50, "1")):
        if percentage >= threshold:
            return grade
    return "0"


def _status_thai(status: str) -> str:
    return {"mastered": "ผ่าน", "developing": "กำลังพัฒนา", "critical": "ต้องเสริม"}[status]


def _html_shell(title: str, body: str, *, extra_css: str = "") -> str:
    return f"""<!doctype html>
<html lang="th"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>{_e(title)}</title><style>
@page {{ size: A4; margin: 12mm; }}
* {{ box-sizing: border-box; }} body {{ margin:0; background:#edf2f7; color:#111827; font-family:"Noto Sans Thai",Tahoma,Arial,sans-serif; font-size:12pt; line-height:1.45; }}
.page {{ width:186mm; min-height:273mm; margin:12px auto; padding:0; background:white; }} .school {{ border-bottom:2px solid #0757d9; padding:0 0 9px; display:flex; justify-content:space-between; gap:12px; }}
.school h1 {{ margin:0; color:#0757d9; font-size:19pt; }} .school p,.muted {{ margin:2px 0; color:#475467; font-size:10pt; }} h2 {{ color:#0757d9; font-size:14pt; margin:13px 0 6px; }} h3 {{ margin:8px 0 4px; font-size:12pt; }}
table {{ width:100%; border-collapse:collapse; margin:7px 0; }} th,td {{ border:1px solid #cbd5e1; padding:5px 6px; vertical-align:top; }} th {{ background:#eaf2ff; color:#12376e; }} .note {{ padding:7px 9px; background:#fff7df; border-left:4px solid #f7ad11; font-size:10pt; }} .badge {{ display:inline-block; padding:2px 7px; border-radius:10px; background:#e8f7f5; color:#087b77; font-size:10pt; }}
.print {{ display:block; margin:12px auto; border:0; border-radius:5px; padding:8px 14px; color:white; background:#0757d9; font:inherit; cursor:pointer; }} .page-break {{ break-before:page; page-break-before:always; }}
{extra_css}
@media print {{ body {{ background:white; }} .page {{ width:auto; min-height:0; margin:0; }} .print {{ display:none !important; }} }}
</style></head><body>{body}<button class="print" onclick="window.print()">พิมพ์เอกสาร</button></body></html>"""


def build_pp5_workbook() -> bytes:
    """Build and validate the three-sheet PP5-style class assessment workbook."""
    wb = Workbook()
    course, scores, summary = wb.active, wb.create_sheet(), wb.create_sheet()
    course.title, scores.title, summary.title = "ข้อมูลรายวิชา", "คะแนนและตัวชี้วัด", "สรุปผล"
    exam, analytics = fixture_exam(), class_analytics()
    attempts = {(a.student_id, a.attempt_number): a for a in seeded_attempts()}

    def title(ws, heading: str, last_col: int) -> None:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        cell = ws.cell(1, 1, heading); cell.font = Font(name="Tahoma", size=18, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE); cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        ws.cell(2, 1, PROTOTYPE_NOTE).font = Font(name="Tahoma", size=11, italic=True, color="667085")
        ws.cell(2, 1).alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A5"; ws.sheet_view.showGridLines = False
        ws.page_setup.orientation, ws.page_setup.paperSize, ws.page_setup.fitToWidth = "landscape", ws.PAPERSIZE_A4, 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.scale = None
        ws.page_margins.left = ws.page_margins.right = 0.25
        ws.print_options.horizontalCentered = True
        ws.print_title_rows = "1:4"

    title(course, "แบบบันทึกข้อมูลรายวิชาและผลการเรียน (ปพ.5 ต้นแบบ)", 6)
    course_rows = [("สถานศึกษา", SCHOOL_NAME), ("รายวิชา", exam.subject), ("ชั้นเรียน", "ม.3/2"), ("หน่วยการเรียนรู้", exam.topic), ("การประเมิน", f"{exam.title} | {exam.duration_minutes} นาที | {exam.total_points} คะแนน"), ("ภาคเรียน/ปีการศึกษา", "1/2569")]
    for row, (label, value) in enumerate(course_rows, 4):
        course.cell(row, 1, label).font = Font(bold=True, color="12376E"); course.cell(row, 1).fill = PatternFill("solid", fgColor=PALE_BLUE)
        course.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6); course.cell(row, 2, value)
    course.freeze_panes = "A4"
    course.column_dimensions["A"].width = 24
    for column in range(2, 7): course.column_dimensions[get_column_letter(column)].width = 20

    title(scores, "บันทึกคะแนนและตัวชี้วัดรายบุคคล", 13)
    headers = ["เลขที่", "รหัสนักเรียน", "ชื่อ-สกุล", "ครั้งที่ 1", "ครั้งที่ 2", "ร้อยละ", "ค 1.1 ม.3/1", "สถานะตัวชี้วัด", "ทักษะที่ต้องเสริม", "ความก้าวหน้า", "ระดับผลการเรียน\n(ประมาณ)", "ลงชื่อครู", "หมายเหตุ"]
    for col, header in enumerate(headers, 1):
        cell = scores.cell(4, col, header); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=TEAL); cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    scores.row_dimensions[4].height = 33
    reports = {student.id: student_report(student.id) for student in seeded_students()}
    for row, student in enumerate(seeded_students(), 5):
        report = reports[student.id]; first, second = attempts[(student.id, 1)], attempts[(student.id, 2)]
        focus_id = report["primary_root_cause"]
        focus = next((node for node in knowledge_nodes() if node.id == focus_id), None)
        values = [
            student.student_number,
            student.student_code,
            student.name,
            first.score,
            second.score,
            report["latest_percentage"],
            focus.indicator if focus else "—",
            "ก้าวหน้า" if focus is None else "ผ่าน" if second.score >= 6 else "ต้องเสริม",
            focus.short_title if focus else "—",
            report["growth"],
            _grade(report["latest_percentage"]),
            "",
            "ต้นแบบ",
        ]
        for col, value in enumerate(values, 1):
            cell = scores.cell(row, col, value); cell.alignment = Alignment(vertical="center", horizontal="center" if col not in (3, 9, 12, 13) else "left", wrap_text=True)
            cell.border = Border(bottom=THIN_GREY)
        if second.score < 6:
            for col in range(1, 14): scores.cell(row, col).fill = PatternFill("solid", fgColor="FFF1F2")
    for col, width in enumerate((8, 14, 28, 10, 10, 10, 16, 16, 20, 12, 14, 18, 14), 1): scores.column_dimensions[get_column_letter(col)].width = width
    scores.auto_filter.ref = f"A4:M{4 + len(reports)}"

    title(summary, "สรุปผลการประเมินระดับชั้น", 7)
    summary_headers = ["รายการ", "ผลการประเมิน", "รายละเอียด", "ผ่าน", "กำลังพัฒนา", "ต้องเสริม", "ข้อเสนอแนะ"]
    for col, value in enumerate(summary_headers, 1):
        cell = summary.cell(4, col, value)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor=TEAL); cell.alignment = Alignment(horizontal="center", wrap_text=True)
    summary_rows = [
        ("จำนวนนักเรียน", 30, "รายชื่อทั้งหมดในชั้น ม.3/2", "", "", "", "ตรวจสอบรายชื่อก่อนบันทึกจริง"),
        ("คะแนนเฉลี่ยครั้งที่ 1", analytics["previous_average"], "ร้อยละ", "", "", "", "ใช้เปรียบเทียบการพัฒนา"),
        ("คะแนนเฉลี่ยครั้งที่ 2", analytics["class_average"], "ร้อยละ", "", "", "", analytics["insight"]["recommendation"]),
    ]
    for item in analytics["node_summary"]:
        dist = item["status_distribution"]
        summary_rows.append((item["node"].short_title, item["average_percentage"], item["node"].indicator, dist["mastered"], dist["developing"], dist["critical"], "จัดกลุ่มซ่อมเสริมตามตัวชี้วัด"))
    for row_index, values in enumerate(summary_rows, 5):
        for col, value in enumerate(values, 1): summary.cell(row_index, col, value)
    for row in summary.iter_rows(min_row=5, max_row=4 + len(summary_rows), max_col=7):
        for cell in row: cell.border = Border(bottom=THIN_GREY); cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col, width in enumerate((24, 17, 24, 12, 16, 14, 34), 1): summary.column_dimensions[get_column_letter(col)].width = width
    summary.freeze_panes = "A5"

    course.print_area = f"A1:{get_column_letter(course.max_column)}{course.max_row}"
    scores.print_area = f"A1:{get_column_letter(scores.max_column)}{scores.max_row}"
    summary.print_area = f"A1:{get_column_letter(summary.max_column)}{summary.max_row}"
    course.page_setup.fitToHeight = 1
    scores.page_setup.fitToHeight = 0
    summary.page_setup.fitToHeight = 1

    for worksheet in (course, scores, summary):
        for row in worksheet.iter_rows():
            for cell in row:
                font = copy(cell.font)
                font.name = "Tahoma"
                cell.font = font

    output = BytesIO(); wb.save(output); data = output.getvalue()
    # A corrupt in-memory workbook must never be returned to the route layer.
    reopened = load_workbook(BytesIO(data), read_only=True)
    if reopened.sheetnames != ["ข้อมูลรายวิชา", "คะแนนและตัวชี้วัด", "สรุปผล"]:
        raise ValueError("PP5 workbook validation failed")
    return data


def build_pa_car_html() -> str:
    """Create a self-contained PA/CAR evidence overview for the seeded class."""
    analytics, exam = class_analytics(), fixture_exam()
    node_rows = "".join(
        f"<tr><td>{_e(item['node'].short_title)}</td>"
        f"<td>{_e(_one_decimal(item['average_percentage']))}%</td>"
        f"<td>{_e(item['status_distribution']['mastered'])}</td>"
        f"<td>{_e(item['status_distribution']['developing'])}</td>"
        f"<td>{_e(item['status_distribution']['critical'])}</td></tr>"
        for item in analytics["node_summary"]
    )
    bars = "".join(
        f"<div class='bar-row'><span>{_e(item['node'].short_title)}</span>"
        f"<div class='track'><i style='width:{max(0.0, min(100.0, float(item['average_percentage']))):.1f}%'></i></div>"
        f"<b>{_e(_one_decimal(item['average_percentage']))}%</b></div>"
        for item in analytics["node_summary"]
    )
    previous_average = _one_decimal(analytics["previous_average"])
    class_average = _one_decimal(analytics["class_average"])
    change = _one_decimal(analytics["class_average"] - analytics["previous_average"])
    body = f"""<main class="page"><header class="school"><div><h1>รายงานหลักฐานผลการพัฒนางาน (PA/CAR)</h1><p>{_e(SCHOOL_NAME)} · กลุ่มสาระการเรียนรู้คณิตศาสตร์</p></div><span class="badge">ชั้น ม.3/2 · 30 คน</span></header>
<p class="note">{_e(PROTOTYPE_NOTE)}</p><table><tr><th>รายวิชา/หน่วย</th><td>{_e(exam.subject)} — {_e(exam.topic)}</td><th>ช่วงประเมิน</th><td>10 มิถุนายน 2569 – 15 กรกฎาคม 2569</td></tr><tr><th>เครื่องมือ</th><td>{_e(exam.title)} 12 ข้อ</td><th>ผู้รับผิดชอบ</th><td>ครูผู้สอน (ต้นแบบ)</td></tr></table>
<h2>1. ผลการเปรียบเทียบก่อนและหลังเรียน</h2><div class="comparison"><section><small>ก่อนเรียน / ครั้งที่ 1</small><strong>{_e(previous_average)}%</strong></section><section><small>หลังเรียน / ครั้งที่ 2</small><strong>{_e(class_average)}%</strong></section><section><small>การเปลี่ยนแปลง</small><strong>+{_e(change)} จุด</strong></section></div>
<p>ผลการประเมินของผู้เรียน 30 คนแสดงให้เห็นว่าคะแนนเฉลี่ยหลังเรียนสูงกว่าก่อนเรียนอย่างชัดเจน โดยยังควรใช้ข้อมูลรายตัวชี้วัดเพื่อวางแผนซ่อมเสริมอย่างต่อเนื่อง</p>
<h2>2. แผนภูมิผลสัมฤทธิ์ตามตัวชี้วัด</h2><div class="chart">{bars}</div><table><thead><tr><th>ตัวชี้วัด/ทักษะ</th><th>เฉลี่ย</th><th>ผ่าน</th><th>กำลังพัฒนา</th><th>ต้องเสริม</th></tr></thead><tbody>{node_rows}</tbody></table>
<h2>3. สรุปและการดำเนินการต่อไป</h2><p><b>ประเด็นเร่งด่วน:</b> {_e(analytics['insight']['headline'])} ({_e(analytics['insight']['focus_node_id'])})</p><p>ครูจะจัดกลุ่มผู้เรียนตามหลักฐานคำตอบผิด ใช้แบบฝึกสั้นทีละขั้น และประเมินซ้ำหลังการสอนเสริม โดยติดตามผู้เรียนที่มีผลต่ำกว่าเกณฑ์เป็นรายบุคคล</p><p class="muted">จัดทำเมื่อ 15 กรกฎาคม 2569 · แหล่งข้อมูล: การประเมินครั้งที่ 1 และ 2 ของชั้น ม.3/2</p></main>"""
    return _html_shell("รายงาน PA/CAR", body, extra_css=".comparison{display:flex;gap:8px}.comparison section{flex:1;padding:10px;border:1px solid #cbd5e1;background:#f8fbff;text-align:center}.comparison strong{display:block;color:#0757d9;font-size:20pt}.chart{border:1px solid #cbd5e1;padding:10px}.bar-row{display:grid;grid-template-columns:90px 1fr 48px;gap:7px;align-items:center;margin:6px 0}.track{height:13px;background:#e5e7eb;border-radius:8px;overflow:hidden}.track i{display:block;height:100%;background:#0e9f9a}@media(max-width:600px){.comparison{display:block}.comparison section{margin-bottom:5px}}")


def build_remedial_html(student_id: str) -> str | None:
    """Create a one-to-two-page individualized remedial worksheet, or None for an unknown learner."""
    report = student_report(student_id)
    if report is None:
        return None
    student = report["student"]
    latest_attempt = report["attempts"][-1]
    advanced = report["primary_root_cause"] is None and report["latest_score"] == latest_attempt.total
    focus_id = report["primary_root_cause"] or "N05"
    node_map = {node.id: node for node in knowledge_nodes()}; focus = node_map[focus_id]
    path = [*prerequisite_path(focus_id), focus_id]
    practices = [q for q in fixture_exam().questions if q.knowledge_node_id == focus_id]
    for question in fixture_exam().questions:
        if question not in practices and len(practices) < 3: practices.append(question)
    practice_html = "".join(
        f"<section class='exercise'><b>ข้อ {_e(index)}.</b> {_e(question.stem)}"
        f"<ol type='A'>{''.join(f'<li>{_e(choice.text)}</li>' for choice in question.choices)}</ol>"
        "<div class='answer-line'>แสดงวิธีทำ: ................................................................................................</div></section>"
        for index, question in enumerate(practices[:3], 1)
    )
    path_html = " → ".join(_e(node_map[node].short_title) for node in path)
    document_heading = "ใบงานต่อยอดรายบุคคล" if advanced else "ใบงานซ่อมเสริมรายบุคคล"
    badge = "ระดับก้าวหน้า" if advanced else "แผนเฉพาะบุคคล"
    target_heading = "เป้าหมายการต่อยอด" if advanced else "เป้าหมายการเรียนรู้"
    target_label = "โจทย์ท้าทาย" if advanced else "จุดเน้น"
    lesson_heading = "แนวคิดต่อยอด" if advanced else "บทเรียนสั้น 5 นาที"
    lesson_steps = (
        "<ol><li>เปรียบเทียบการแยกตัวประกอบกับสูตรกำลังสอง</li><li>อธิบายเหตุผลและตรวจคำตอบด้วยอีกวิธีหนึ่ง</li><li>สร้างโจทย์ใหม่ที่มีรากเป็นเศษส่วนหรือมีพารามิเตอร์</li></ol>"
        if advanced else
        "<ol><li>จัดรูปสมการให้เท่ากับศูนย์</li><li>แยกตัวประกอบโดยตรวจเครื่องหมายอย่างเป็นระบบ</li><li>ใช้สมบัติผลคูณเป็นศูนย์ แล้วแทนค่าตรวจคำตอบ</li></ol>"
    )
    exercise_heading = "แบบฝึกต่อยอด (3 ข้อ)" if advanced else "แบบฝึกเน้นจุด (3 ข้อ)"
    observation = (
        "□ อธิบายได้มากกว่า 1 วิธี &nbsp; □ ให้เหตุผลเลือกวิธีได้ &nbsp; □ สร้างโจทย์ท้าทายได้"
        if advanced else
        "□ อ่านโจทย์และจัดรูปได้ &nbsp; □ แยกตัวประกอบได้ &nbsp; □ ตรวจคำตอบได้"
    )
    follow_up = (
        "ครูชวนสะท้อนกลยุทธ์และมอบโจทย์เปิดเพื่อเก็บหลักฐานการคิดขั้นสูงในคาบถัดไป"
        if advanced else
        "ครูติดตามผลหลังทำใบงานและให้ทำแบบประเมินซ้ำสั้น ๆ ในคาบถัดไป"
    )
    body = f"""<main class="page" data-focus-node="{_e(focus_id)}"><header class="school"><div><h1>{document_heading}</h1><p>{_e(SCHOOL_NAME)} · คณิตศาสตร์ ม.3</p></div><span class="badge">{badge}</span></header><p class="note">{_e(PROTOTYPE_NOTE)}</p>
<table><tr><th>ชื่อ-สกุล</th><td>{_e(student.name)}</td><th>เลขที่</th><td>{_e(student.student_number)}</td></tr><tr><th>ห้อง</th><td>{_e(student.room)}</td><th>ผลล่าสุด</th><td>{_e(report['latest_score'])}/{_e(latest_attempt.total)} ({_e(_one_decimal(report['latest_percentage']))}%)</td></tr></table>
<h2>{target_heading}</h2><p>{target_label}: <b>{_e(focus.title)}</b> ({_e(focus_id)} · {_e(focus.indicator)})</p><p><b>เส้นทางความรู้:</b> {path_html}</p><div class="lesson"><h3>{lesson_heading}</h3><p>{_e(report['diagnosis'])}</p>{lesson_steps}</div>
<h2>{exercise_heading}</h2>{practice_html}<h2>{'การสะท้อนและติดตาม' if advanced else 'การติดตามโดยครู'}</h2><table><tr><th>สิ่งที่สังเกต</th><td>{observation}</td></tr><tr><th>นัดติดตาม</th><td>............................................................</td></tr><tr><th>บันทึกครู</th><td>................................................................................................................</td></tr></table><p class="muted">{follow_up}</p></main>"""
    return _html_shell("ใบงานต่อยอด" if advanced else "ใบงานซ่อมเสริม", body, extra_css=".lesson{border:1px solid #9edbd7;background:#f2fcfb;padding:8px 11px}.exercise{border:1px solid #cbd5e1;padding:8px 10px;margin:7px 0;break-inside:avoid}.exercise ol{margin:4px 0 5px;padding-left:26px}.answer-line{color:#667085;font-size:10pt}")


def build_exam_html() -> str:
    """Create a print-ready exam plus a separate teacher answer-key page."""
    exam = fixture_exam()
    questions = "".join(f"<section class='question'><b>{_e(index)}. </b>{_e(question.stem)}<div class='choices'>{''.join(f'<div>{_e(choice.id)}. {_e(choice.text)}</div>' for choice in question.choices)}</div></section>" for index, question in enumerate(exam.questions, 1))
    key = "".join(f"<tr><td>{_e(index)}</td><td>{_e(question.correct_choice_id)}</td><td>{_e(question.solution_explanation)}</td></tr>" for index, question in enumerate(exam.questions, 1))
    body = f"""<main class="page"><header class="school"><div><h1>{_e(SCHOOL_NAME)}</h1><p>แบบทดสอบปลายหน่วยการเรียนรู้</p></div><span class="badge">คณิตศาสตร์ ม.3</span></header><div class="exam-title"><h2>{_e(exam.title)}</h2><p>เรื่อง {_e(exam.topic)} · เวลา {_e(exam.duration_minutes)} นาที · คะแนนเต็ม {_e(exam.total_points)} คะแนน</p></div><div class="fields">ชื่อ-สกุล ........................................................................ เลขที่ ............ ห้อง ............</div><h3>คำชี้แจง</h3><ol class="instructions"><li>เลือกคำตอบที่ถูกต้องที่สุดเพียงข้อเดียว</li><li>ทำเครื่องหมายคำตอบลงในกระดาษคำตอบ และแสดงวิธีคิดเมื่อครูกำหนด</li><li>ตรวจสอบชื่อ เลขที่ และห้องเรียนก่อนส่งข้อสอบ</li></ol><div class="questions">{questions}</div></main>
<main class="page page-break answer-key"><header class="school"><div><h1>เฉลยสำหรับครู</h1><p>{_e(SCHOOL_NAME)} · {_e(exam.title)}</p></div><span class="badge">ห้ามแจกนักเรียน</span></header><p class="note">{_e(PROTOTYPE_NOTE)}</p><table><thead><tr><th>ข้อ</th><th>คำตอบ</th><th>แนวคิด</th></tr></thead><tbody>{key}</tbody></table></main>"""
    return _html_shell("แบบทดสอบคณิตศาสตร์ ม.3", body, extra_css=".exam-title{text-align:center;border-bottom:1px solid #cbd5e1;margin:8px 0}.exam-title h2{margin:4px}.fields{padding:8px 0;border-bottom:1px solid #cbd5e1}.instructions{margin:4px 0 8px;padding-left:25px}.questions{column-count:2;column-gap:13mm}.question{break-inside:avoid;border-bottom:1px dotted #94a3b8;padding:6px 0}.choices{padding-left:17px;margin-top:3px}.answer-key table{font-size:11pt}@media(max-width:650px){.questions{column-count:1}}")
