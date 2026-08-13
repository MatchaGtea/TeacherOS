from io import BytesIO

from openpyxl import load_workbook

from backend.app.exports import build_exam_html, build_pa_car_html, build_pp5_workbook, build_remedial_html
from backend.app.exports import documents as export_documents
from backend.app.domain.services import fixture_exam


def test_pp5_workbook_has_required_sheets_and_all_learners():
    workbook = load_workbook(BytesIO(build_pp5_workbook()), data_only=True)
    assert workbook.sheetnames == ["ข้อมูลรายวิชา", "คะแนนและตัวชี้วัด", "สรุปผล"]
    scores = workbook["คะแนนและตัวชี้วัด"]
    assert scores.max_row == 34
    assert scores["C5"].value == "นายสมชาย ใจดี"
    assert scores["D5"].value == 5 and scores["E5"].value == 8
    assert scores["F5"].value == 66.7 and scores["G5"].value == "ค 1.1 ม.3/1"
    assert "เอกสารต้นแบบ" in workbook["ข้อมูลรายวิชา"]["A2"].value


def test_pp5_workbook_handles_perfect_students_and_has_print_configuration():
    workbook = load_workbook(BytesIO(build_pp5_workbook()), data_only=True)
    scores = workbook["คะแนนและตัวชี้วัด"]
    assert scores["C7"].value == "นายเด็กดี"
    assert scores["G7"].value == "—"
    assert scores["H7"].value == "ก้าวหน้า"
    assert scores["I7"].value == "—"

    expected = {
        "ข้อมูลรายวิชา": ("A4", "$A$1:$F$9"),
        "คะแนนและตัวชี้วัด": ("A5", "$A$1:$M$34"),
        "สรุปผล": ("A5", "$A$1:$G$12"),
    }
    for sheet_name, (freeze_panes, area) in expected.items():
        sheet = workbook[sheet_name]
        assert sheet.freeze_panes == freeze_panes
        assert area in str(sheet.print_area)
        assert sheet.print_title_rows == "$1:$4"
        assert sheet.page_setup.fitToWidth == 1
        assert sheet.sheet_properties.pageSetUpPr.fitToPage is True


def test_html_exports_are_complete_and_print_ready():
    pa_car, exam, remedial = build_pa_car_html(), build_exam_html(), build_remedial_html("STU001")
    for document in (pa_car, exam, remedial):
        assert document and "<meta charset=\"utf-8\">" in document
        assert "@page" in document and "@media print" in document and "window.print()" in document
        assert "เอกสารต้นแบบ" in document
    assert "30 คน" in pa_car and "ก่อนเรียน" in pa_car and "หลังเรียน" in pa_car
    assert "column-count:2" in exam and "เฉลยสำหรับครู" in exam and exam.count("class='question'") == 12
    assert "นายสมชาย ใจดี" in remedial and "แบบฝึกเน้นจุด (3 ข้อ)" in remedial
    assert remedial.count("class='exercise'") == 3


def test_perfect_student_gets_advanced_challenge_document():
    document = build_remedial_html("STU003")
    assert document is not None
    assert "ใบงานต่อยอดรายบุคคล" in document
    assert "ระดับก้าวหน้า" in document and 'data-focus-node="N05"' in document
    assert "แบบฝึกต่อยอด (3 ข้อ)" in document
    assert document.count("class='exercise'") == 3
    assert "ใบงานซ่อมเสริม" not in document and "ควรทบทวน" not in document


def test_html_has_title_thai_language_and_no_external_runtime_assets():
    documents = (build_pa_car_html(), build_exam_html(), build_remedial_html("STU001"))
    for document in documents:
        assert document is not None
        lowered = document.lower()
        assert '<html lang="th">' in lowered
        assert "<title>" in lowered and "</title>" in lowered
        assert "http://" not in lowered and "https://" not in lowered
        assert "<link" not in lowered and "<script src" not in lowered
        assert "@import" not in lowered and "url(" not in lowered


def test_exam_escapes_dynamic_content(monkeypatch):
    tainted = fixture_exam().model_copy(update={"title": '<img src="x" onerror="bad()">'})
    monkeypatch.setattr(export_documents, "fixture_exam", lambda: tainted)
    document = build_exam_html()
    assert '<img src="x" onerror="bad()">' not in document
    assert "&lt;img src=&quot;x&quot; onerror=&quot;bad()&quot;&gt;" in document


def test_remedial_is_none_for_unknown_student():
    assert build_remedial_html("MISSING") is None
