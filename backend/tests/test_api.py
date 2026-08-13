from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app.main import app


client = TestClient(app)


def test_health_and_demo_exam_contract():
    assert client.get("/api/health").json() == {"status": "ok"}
    response = client.get("/api/exams/demo")
    assert response.status_code == 200
    payload = response.json()
    assert len(payload["questions"]) == 12
    assert payload["pipeline"]["sympy_verified_count"] == 12
    assert all(question["sympy_verified"] for question in payload["questions"])


def test_fixture_generation_and_scoring_do_not_change_class_dataset():
    before = client.get("/api/classes/demo/analytics").json()
    generated = client.post(
        "/api/exams/generate",
        json={
            "grade": "ม.3",
            "subject": "คณิตศาสตร์",
            "topic": "สมการกำลังสองตัวแปรเดียว",
            "question_count": 12,
            "difficulty": "mixed",
            "mode": "fixture",
        },
    )
    assert generated.status_code == 200
    exam = generated.json()
    answers = {question["id"]: question["correct_choice_id"] for question in exam["questions"]}
    result = client.post(
        "/api/attempts/score",
        json={"student_name": "ผู้เรียนทดลอง", "exam_id": exam["id"], "answers": answers},
    )
    assert result.status_code == 200
    assert result.json()["score"] == 12
    assert result.json()["percentage"] == 100
    assert client.get("/api/classes/demo/analytics").json() == before


def test_reports_and_unknown_student_http_contract():
    report = client.get("/api/students/STU001/report")
    assert report.status_code == 200
    assert report.json()["student"]["name"] == "นายสมชาย ใจดี"
    assert client.get("/api/students/MISSING/report").status_code == 404
    assert client.get("/api/students/MISSING/remedial").status_code == 404


def test_html_and_workbook_export_routes():
    for path, marker in (
        ("/api/exports/exam", "เฉลยสำหรับครู"),
        ("/api/students/STU001/remedial", "นายสมชาย ใจดี"),
        ("/api/exports/pa-car", "PA/CAR"),
    ):
        response = client.get(path)
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("text/html")
        assert marker in response.text

    response = client.get("/api/exports/pp5.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert workbook.sheetnames == ["ข้อมูลรายวิชา", "คะแนนและตัวชี้วัด", "สรุปผล"]
